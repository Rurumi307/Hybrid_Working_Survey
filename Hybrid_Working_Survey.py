import pandas as pd
import datetime
from openpyxl.utils import get_column_letter
from Hybrid_Working_module import *

class DataProcessor:
    def __init__(self, year, month, national_holidays, supervisor_list, last_month_count):
        """
        Initialize the DataProcessor class.

        Args:
            year (int): The year.
            month (int): The month.
            national_holidays (list): List of national holidays.
            supervisor_list (list): List of supervisors.
            last_month_count (int): The count of people from the last month.
        """
        self.year = year
        self.month = month
        self.YYYY_MM = f'{self.year}_{str(self.month).zfill(2)}'
        self.file_path = f'./output_{self.year}_{str(self.month).zfill(2)}.xlsx'
        self.national_holidays = national_holidays
        self.df_it = pd.read_excel('./IT.xlsx')
        self.count_people = len(self.df_it) + 2
        self.df_it_talk = pd.read_excel('./IT.xlsx', sheet_name='one_on_one_talk_statistic')
        self.max_row = len(self.df_it_talk)
        self.talk_sheet_name = f'one-on-one_TalkStatistic_{str(month).zfill(2)}{year}'
        self.supervisor_list = supervisor_list
        self.last_month_count = last_month_count

    def _process_data(self):
        """
        Process the data by creating a working days list, joining dataframes, and writing to Excel.
        """
        # Create a list of working days and a list of all days in a given month
        self.working_days_list, days_list = create_working_days_list(self.year, self.month, self.national_holidays)
        # Create dataframes for working days and dates
        self.df_work_day = pd.DataFrame(self.working_days_list) 
        self.df_date = pd.DataFrame({'Date': days_list})

        # Join the working days and dates dataframes on the 'Date' column
        self.df_date = self.df_date.join(self.df_work_day.set_index('Date'), on='Date')
        # Transpose the working days dataframe and reorder the columns
        self.df_update = self.df_work_day.set_index('Week').T
        order_list = ['Date'] + list(range(0, 69))
        self.df_update = pd.concat([self.df_it, self.df_update], axis=1).reindex(order_list)
        
        # Write the working days dataframe to Excel
        write_dataframe_to_excel(self.df_update, self.file_path, self.YYYY_MM)
        return self

    def _create_excel_func(self):
        """
        Create Excel functions and write the dataframe to Excel.

        This function creates Excel functions and writes the dataframe to an Excel file. It iterates over the data and performs
        calculations based on the values in each row. The calculated values are then appended to the data.

        Returns:
            self: The instance of the class.
        """
        data = self.df_date.values.tolist()
        column_index = 0
        for idx, row in enumerate(data):
            column_letter = get_column_letter(6 + column_index)
            cell_range = f'${column_letter}$3:${column_letter}${self.count_people}'
            it_total_count_formula = f'''=COUNTA('{self.YYYY_MM}'!$E$3:$E${self.count_people})'''
            if isinstance(row[1], str):
                cnt_wfh, pct_wfh, cnt_office_work, pct_office_work, cnt_leave, pct_leave, \
                cnt_empty, pct_empty, cnt_filled, pct_filled, check_cnt = calculate_metrics(self.YYYY_MM, cell_range, idx)
                data[idx].extend([
                    cnt_wfh, pct_wfh, cnt_office_work, pct_office_work, cnt_leave, pct_leave,
                    cnt_empty, pct_empty, cnt_filled, pct_filled, it_total_count_formula, check_cnt
                ])
                column_index += 1
            else:
                data[idx].extend(['-' for _ in range(10)])
                data[idx].append(it_total_count_formula)
        self.df_date = pd.DataFrame(data, columns=['Date', 'Week', '居家工作', '居家工作%', '進公司', '進公司%', '請假', '請假%', '未填', '未填%', '已填', '已填%', 'IT總人數', 'Column2'])
        self.df_date = self.df_date.drop('Week', axis=1)
        write_dataframe_to_excel(self.df_date, self.file_path, f'{self.YYYY_MM}_Func')
        return self
    
    def _update_it_talk_data(self):
        """
        Update the IT talk data by adding a new dataframe and writing to Excel.
        """
        self.df_work_day.drop('Week', axis=1, inplace=True)
        self.df_work_day = self.df_work_day.set_index('Date').T
        self.df_it_talk = pd.concat([self.df_it_talk, self.df_work_day], axis=1)
        write_dataframe_to_excel(self.df_it_talk, self.file_path, self.talk_sheet_name)
        return self

    def _process_talk_data(self):
        """
        Process the talk data by adding new columns, calculating metrics, and writing to Excel.
        """
        _sheet_name = 'one-on-one_Talk_月結'
        self.df_talk_func = self.df_it_talk[~self.df_it_talk['上級主管'].isnull()].copy()
        self.df_talk_func['YearMonth'] = datetime.datetime(self.year, self.month, 1).strftime("%b %Y")
        self.df_talk_func = self.df_talk_func[['YearMonth', '工號', '姓名', '事業處名', '處級名', '部級名', '上級主管']]
        cnt_working_days = len(create_working_days_list(self.year, self.month, self.national_holidays)[0])
        data = self.df_talk_func.values.tolist()
        for idx, row in enumerate(data):
            column_letter = get_column_letter(6 + cnt_working_days)
            cell_range = f'G${idx+4}:{column_letter}{idx+4}'
            activity_monthly_count = f'''=COUNTIFS('{self.talk_sheet_name}'!{cell_range},"v")'''
            achieve_goals = f'''=IF(H{idx+2}>0, "OK", "need to arrange")'''
            data[idx].extend([activity_monthly_count, achieve_goals])
        self.df_talk_func = pd.DataFrame(data, columns=['YearMonth', '員工工號', '員工姓名', '事業處名', '處級名', '部級名', '上級主管',
                                                   'one-on-one activity_monthly count', 'Achieve goals'])
        write_dataframe_to_excel(self.df_talk_func, self.file_path, _sheet_name)
        return self
    
    def _process_supervisor_data(self):
        """
        Process the supervisor data by adding new rows, calculating metrics, and writing to Excel.
        """
        _sheet_name = 'one-on-one_Talk_月分析'
        supervisor_data = []
        year_month = datetime.datetime(self.year, self.month, 1).strftime("%b %Y")
        
        for supervisor in self.supervisor_list:
            supervisor_talked_cnt = f'''=COUNTIFS('one-on-one_Talk_月結'!$G$1:$G${self.max_row},"{supervisor}",'one-on-one_Talk_月結'!$I$1:$I${self.last_month_count},"OK")'''
            supervisor_member_cnt = f'''=COUNTIFS('one-on-one_Talk_月結'!$G$1:$G${self.max_row},"{supervisor}")'''
            supervisor_data.append([year_month, supervisor, supervisor_talked_cnt, supervisor_member_cnt])
        
        supervisor_data.append(['', '', f'=SUM(D2:D9)', f'=SUM(E2:E9)'])
        supervisor_data.append([])
        
        for supervisor in self.supervisor_list:
            supervisor_talked_cnt = f'''=COUNTIFS('one-on-one_Talk_月結'!$G${1+self.max_row}:$G${self.max_row+self.last_month_count},"{supervisor}",'one-on-one_Talk_月結'!$I${1+self.last_month_count}:$I${self.max_row+self.last_month_count},"OK")'''
            supervisor_member_cnt = f'''=COUNTIFS('one-on-one_Talk_月結'!$G$1:$G${self.max_row+self.last_month_count},"{supervisor}")'''
            supervisor_data.append(['', supervisor, supervisor_talked_cnt, supervisor_member_cnt])
        
        df_analysis = pd.DataFrame(supervisor_data, columns=['YearMonth', '上級主管', 'Talked', 'member count'])
        write_dataframe_to_excel(df_analysis, self.file_path, _sheet_name)

    def process(self):
        """
        Process the data by calling the necessary methods.
        """
        self._process_data()
        self._create_excel_func()
        merge_excel_cells(self.file_path, self.YYYY_MM)
        self._update_it_talk_data()
        self._process_talk_data()
        self._process_supervisor_data()