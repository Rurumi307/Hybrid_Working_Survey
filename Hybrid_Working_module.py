
import os 
import calendar
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# setting Sunday as the first day of the week
calendar.setfirstweekday(6)
def get_week_of_month(year, month, day):
    """
    Get the week of the month for a given date.

    Args:
        year (int): The year of the date.
        month (int): The month of the date.
        day (int): The day of the date.

    Returns:
        str: The week of the month in Chinese characters.

    Examples:
        >>> get_week_of_month(2022, 12, 31)
        '第五周'
    """
    x = np.array(calendar.monthcalendar(year, month))
    week_of_month = np.where(x == day)[0][0] + 1
    chinese_weeks = {
        1: "第一周",
        2: "第二周",
        3: "第三周",
        4: "第四周",
        5: "第五周"
    }
    return chinese_weeks.get(week_of_month, "")

def create_working_days_list(year, month, national_holidays):
    """
    Create a list of working days and a list of all days in a given month.

    Args:
        year (int): The year.
        month (int): The month.
        national_holidays (list): List of national holidays in the format "YYYY/MM/DD".

    Returns:
        tuple: A tuple containing the working days list and the days list.
    """
    # Get the number of days in the month
    _, days = calendar.monthrange(year, month)

    working_days_list = [
        {
            "Date": f"{year}/{month}/{day}",
            "Week": get_week_of_month(year, month, day)
        }
        for day in range(1, days + 1)
        if calendar.weekday(year, month, day) < 5
    ]

    days_list = [f"{year}/{month}/{day}" for day in range(1, days + 1)]

    return working_days_list, days_list
        
def write_dataframe_to_excel(dataframe, file_path, sheet_name):
    """
    Write a pandas DataFrame to an Excel file.

    Args:
        dataframe (pd.DataFrame): The DataFrame to be written.
        file_path (str): The path of the Excel file.
        sheet_name (str): The name of the sheet in the Excel file.

    Returns:
        None
    """
    if not os.path.isfile(file_path):
        # If the file does not exist, create a new Excel file and write the DataFrame to it
        with pd.ExcelWriter(file_path, mode='w', engine='openpyxl') as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # If the file already exists, append the DataFrame to the existing file
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists="replace") as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

def merge_excel_cells(file_path, sheet_name):
    """
    Merge cells with the same value in each row of the specified sheet in the Excel file.

    Args:
        file_path (str): The path of the Excel file.
        sheet_name (str): The name of the sheet.

    Returns:
        None
    """
    # Load the workbook
    book = load_workbook(file_path)
    sheet = book[sheet_name]

    # Iterate through each row
    for row in sheet.iter_rows():
        prev_value = None
        merge_start = None

        # Iterate through each cell in the row
        for cell in row:
            # Check if the cell value is different from the previous value
            if cell.value != prev_value:
                # Check if there is a merge in progress
                if merge_start is not None:
                    merge_end = cell.offset(column=-1)
                    # Merge the cells
                    sheet.merge_cells(start_row=merge_start.row, start_column=merge_start.column,
                                      end_row=merge_end.row, end_column=merge_end.column)
                merge_start = cell
            prev_value = cell.value

        # Check if there is a merge in progress at the end of the row
        if merge_start is not None:
            merge_end = row[-1]
            # Merge the cells
            sheet.merge_cells(start_row=merge_start.row, start_column=merge_start.column,
                              end_row=merge_end.row, end_column=merge_end.column)

    # Save the workbook
    book.save(file_path)

def calculate_metrics(sheet_name, cell_range, index):
    """
    Calculates various metrics based on the given sheet name, cell range, and index.

    Args:
        sheet_name (str): The name of the sheet containing the data.
        cell_range (str): The range of cells to consider for calculations.
        index (int): The index used for referencing cells in the calculations.

    Returns:
        tuple: A tuple containing the following metrics:
            - count_wfh (str): The count of employees working from home.
            - percentage_wfh (str): The percentage of employees working from home.
            - count_office_work (str): The count of employees working in the office.
            - percentage_office_work (str): The percentage of employees working in the office.
            - count_leave (str): The count of employees on leave.
            - percentage_leave (str): The percentage of employees on leave.
            - count_empty (str): The count of cells that are not filled.
            - percentage_empty (str): The percentage of cells that are not filled.
            - count_filled (str): The count of cells that are filled.
            - percentage_filled (str): The percentage of cells that are filled.
            - check_count (str): A check to verify if the total count matches the sum of individual counts.
    """
    # Calculate the count and percentage of employees working from home
    count_wfh = f'''=COUNTIF('{sheet_name}'!{cell_range},"居家工作")'''
    percentage_wfh = f'''ROUND(([@居家工作]/[@IT總人數])*100,2) & "%"'''

    # Calculate the count and percentage of employees working in the office
    count_office_work = f'''=COUNTIF('{sheet_name}'!{cell_range},"進公司") + COUNTIF('{sheet_name}'!{cell_range},"出差")'''
    percentage_office_work = f'''ROUND(([@進公司]/[@IT總人數])*100,2) & "%"'''

    # Calculate the count and percentage of employees on leave
    count_leave = f'''=COUNTIF('{sheet_name}'!{cell_range},"*假")'''
    percentage_leave = f'''ROUND(([@請假]/[@IT總人數])*100,2) & "%"'''

    # Calculate the count and percentage of empty cells
    count_empty = f'''=COUNTIF('{sheet_name}'!{cell_range},"")'''
    percentage_empty = f'''ROUND(([@未填]/[@IT總人數])*100,2) & "%"'''

    # Calculate the count and percentage of filled cells
    count_filled = f'''=B{index+2}+D{index+2}+F{index+2}'''
    percentage_filled = f'''ROUND(([@已填]/[@IT總人數])*100,2) & "%"'''

    # Check if the total count matches the sum of individual counts
    check_count = f'''=IF(($H{index+2}+$J{index+2})=$L{index+2},"OK", "error")'''

    return count_wfh, percentage_wfh, count_office_work, percentage_office_work, count_leave, percentage_leave, count_empty, percentage_empty, count_filled, percentage_filled, check_count
